import streamlit as st
import gspread
import json
from datetime import datetime
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- KONFIGURASI HALAMAN WEB ---
st.set_page_config(page_title="CASHBOOK-PRO_v1", page_icon="📱", layout="centered")

# ==================== SISTEM KEAMANAN LOGIN ====================
def cek_login():
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if not st.session_state["authenticated"]:
        st.subheader("🔒 Akses Terbatas - Cashbook Kantor")
        password = st.text_input("Masukkan Password Akses:", type="password")
        if st.button("Masuk", type="primary", use_container_width=True):
            if password == "kantor123": 
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("❌ Password salah!")
        return False
    return True

if cek_login():
    # ==================== KONEKSI GOOGLE SHEETS VIA GSPREAD ====================
    def dapatkan_koneksi_spreadsheet():
        try:
            kredensial_mentah = st.secrets["connections"]["gsheets"]["service_account"]
            spreadsheet_url = st.secrets["connections"]["gsheets"]["spreadsheet"]
            kredensial_json = json.loads(kredensial_mentah, strict=False)
            
            gc = gspread.service_account_from_dict(kredensial_json)
            sh = gc.open_by_url(spreadsheet_url)
            return sh
        except Exception as e:
            st.error(f"❌ Masalah Koneksi Google Sheets: {e}")
            st.stop()

    # ==================== FUNGSI OTOMATISASI VISUAL SHEETS ====================
    def perbarui_desain_visual_sheet(ws):
        try:
            ws.columns_auto_resize(1, 7)
            total_baris = len(ws.get_all_values())
            if total_baris == 0:
                return

            requests = []
            
            # Format Header (Baris 1): Latar Biru Tua, Teks Putih Tebal, Rata Tengah
            requests.append({
                "repeatCell": {
                    "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 7},
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.1, "green": 0.4, "blue": 0.8},
                            "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True},
                            "horizontalAlignment": "CENTER"
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                }
            })
            
            # Format Efek Zebra untuk Baris Data
            for i in range(1, total_baris):
                warna_latar = {"red": 0.92, "green": 0.96, "blue": 1.0} if i % 2 == 1 else {"red": 1.0, "green": 1.0, "blue": 1.0}
                requests.append({
                    "repeatCell": {
                        "range": {"sheetId": ws.id, "startRowIndex": i, "endRowIndex": i + 1, "startColumnIndex": 0, "endColumnIndex": 7},
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": warna_latar,
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment)"
                    }
                })
            
            ws.spreadsheet.batch_update({"requests": requests})
        except Exception:
            pass

    KAMUS_HARI = {
        "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu",
        "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu", "Sunday": "Minggu"
    }

    def ambil_hari_ini():
        hari_inggris = datetime.now().strftime("%A")
        return KAMUS_HARI.get(hari_inggris, hari_inggris)

    # Ambil Dokumen Spreadsheet Utama
    sh = dapatkan_koneksi_spreadsheet()
    daftar_sheet = [ws.title for ws in sh.worksheets()]
    
    BULAN_INDO = {
        "January": "Januari", "February": "Februari", "March": "Maret", "April": "April",
        "May": "Mei", "June": "Juni", "July": "Juli", "August": "Agustus",
        "September": "September", "October": "Oktober", "November": "November", "December": "Desember"
    }
    bulan_inggris = datetime.now().strftime("%B")
    bulan_lokal = BULAN_INDO.get(bulan_inggris, bulan_inggris)
    tahun_sekarang = datetime.now().strftime("%Y")
    rekomendasi_sheet_baru = f"{bulan_lokal}_{tahun_sekarang}"

    # --- Pilihan Dropdown Sheet Aktif di Navigasi Atas ---
    st.title("📱 Cashbook Kantor (Cloud)")
    sheet_aktif = st.selectbox("📂 Pilih Tab / Bulan Kerja Saat Ini:", options=daftar_sheet)
    st.write("---")

    # Membaca data dari sheet yang dipilih
    ws_aktif = sh.worksheet(sheet_aktif)
    semua_data = ws_aktif.get_all_records()

    if len(semua_data) == 0:
        df_master = pd.DataFrame(columns=["Tanggal", "Hari", "Deskripsi", "Nota", "Debit", "Kredit", "Saldo"])
    else:
        df_master = pd.DataFrame(semua_data)
        df_master = df_master[df_master["Tanggal"] != "Tanggal"]

    if not df_master.empty and "Saldo" in df_master.columns:
        df_master["Debit"] = pd.to_numeric(df_master["Debit"], errors='coerce').fillna(0).astype(int)
        df_master["Kredit"] = pd.to_numeric(df_master["Kredit"], errors='coerce').fillna(0).astype(int)
        df_master["Saldo"] = pd.to_numeric(df_master["Saldo"], errors='coerce').fillna(0).astype(int)
        saldo_terakhir_aktif = int(df_master.iloc[-1]["Saldo"])
        total_pengeluaran = int(df_master["Debit"].sum())
        total_pemasukan = int(df_master["Kredit"].sum())
    else:
        saldo_terakhir_aktif = 0
        total_pengeluaran = 0
        total_pemasukan = 0

    st.metric(label=f"Sisa Saldo Kas di Tab ({sheet_aktif})", value=f"Rp {saldo_terakhir_aktif:,}")
    col_met1, col_met2 = st.columns(2)
    col_met1.metric(label="Total Keluar Tab Ini", value=f"Rp {total_pengeluaran:,}")
    col_met2.metric(label="Total Masuk Tab Ini", value=f"Rp {total_pemasukan:,}")
    st.write("---")

    # ==================== FITUR TAMBAH SHEET BARU (ESTAFET SALDO) ====================
    with st.expander("➕ Menu Tambah Sheet / Tab Baru"):
        st.write("Fitur ini akan membuat tab bulanan baru di Google Sheets Anda dan mengestafetkan sisa saldo terakhir.")
        nama_baru = st.text_input("Nama Sheet Baru:", value=rekomendasi_sheet_baru)
        
        if st.button("🚀 Buat dan Pindahkan Saldo Terakhir", use_container_width=True):
            if nama_baru in daftar_sheet:
                st.warning(f"⚠️ Tab '{nama_baru}' sudah ada di Google Sheets Anda!")
            else:
                try:
                    ws_baru = sh.add_worksheet(title=nama_baru, rows="1000", cols="20")
                    ws_baru.append_row(["Tanggal", "Hari", "Deskripsi", "Nota", "Debit", "Kredit", "Saldo"])
                    baris_awal = [
                        datetime.now().strftime("%Y-%m-%d"),
                        ambil_hari_ini(),
                        f"Estafet Saldo Akhir dari {sheet_aktif}",
                        "-",
                        0,
                        0,
                        int(saldo_terakhir_aktif)
                    ]
                    ws_baru.append_row(baris_awal)
                    perbarui_desain_visual_sheet(ws_baru)
                    st.success(f"🎉 Tab '{nama_baru}' berhasil dibuat dengan saldo awal Rp {saldo_terakhir_aktif:,}!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal membuat sheet baru: {e}")

    st.write("---")

    # ==================== FORM INPUT TRANSAKSI HARIAN ====================
    with st.form("form_transaksi_gspread", clear_on_submit=True):
        st.subheader(f"📝 Tambah Data Ke Tab: {sheet_aktif}")
        deskripsi = st.text_input("Deskripsi / Keperluan:")
        tipe = st.selectbox("Jenis Transaksi:", options=["Kredit (Pemasukan / Uang Masuk)", "Debit (Pengeluaran)"])
        jumlah = st.number_input("Jumlah Uang (Rp):", min_value=0, value=0, step=5000)
        cb_nota = st.checkbox("Ada Bukti Nota Fisik")
        
        if st.form_submit_button("Simpan ke Google Sheets", type="primary", use_container_width=True):
            if deskripsi and jumlah > 0:
                status_nota = "Ada" if cb_nota else "Tidak Ada"
                debit = jumlah if "Debit" in tipe else 0
                kredit = jumlah if "Kredit" in tipe else 0
                
                saldo_baru = (saldo_terakhir_aktif - debit) if debit > 0 else (saldo_terakhir_aktif + kredit)
                
                baris_data = [
                    datetime.now().strftime("%Y-%m-%d"),
                    ambil_hari_ini(),
                    deskripsi,
                    status_nota,
                    int(debit),
                    int(kredit),
                    int(saldo_baru)
                ]
                
                try:
                    ws_aktif.append_row(baris_data)
                    perbarui_desain_visual_sheet(ws_aktif)
                    st.success(f"Berhasil disimpan ke Tab {sheet_aktif}!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal menulis data: {e}")
            else:
                st.warning("⚠️ Mohon isi deskripsi dan nominal uang dengan benar!")

    # ==================== TABEL DATA SINKRON (STREAMLIT APP TAMPILAN) ====================
    if not df_master.empty:
        st.write("---")
        st.subheader(f"📊 Tabel Data Terkini - Tab {sheet_aktif}")
        
        df_tampil = df_master.copy()
        df_tampil["Debit"] = df_tampil["Debit"].apply(lambda x: f"Rp {x:,}")
        df_tampil["Kredit"] = df_tampil["Kredit"].apply(lambda x: f"Rp {x:,}")
        df_tampil["Saldo"] = df_tampil["Saldo"].apply(lambda x: f"Rp {x:,}")
        
        st.markdown(
            """
            <style>
            .dataframe th { text-align: center !important; }
            .dataframe td { text-align: center !important; }
            </style>
            """, 
            unsafe_allow_html=True
        )
        
        st.dataframe(df_tampil, use_container_width=True, hide_index=True)

        # ==================== FITUR DOWNLOAD SINKRON VISUAL EXCEL ====================
        st.write("---")
        st.subheader("📥 Download File Excel Tab Ini")
        
        if st.button("🔍 Siapkan File Excel untuk Diunduh", use_container_width=True):
            buffer = io.BytesIO()
            
            # Buat file excel mentah terlebih dahulu
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_master.to_excel(writer, index=False, sheet_name=sheet_aktif)
                
                # Mengambil objek lembar kerja aktif dari openpyxl untuk mendesain file download
                workbook = writer.book
                worksheet = workbook[sheet_aktif]
                
                # Definisikan warna kosmetik yang persis dengan Google Sheets Anda
                warna_header = PatternFill(start_color="1A66CC", end_color="1A66CC", fill_type="solid") # Biru Tua
                warna_zebra = PatternFill(start_color="EBF5FF", end_color="EBF5FF", fill_type="solid")  # Biru Muda Transparan
                teks_putih = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                teks_biasa = Font(name="Calibri", size=11, bold=False, color="000000")
                rata_tengah = Alignment(horizontal="center", vertical="center")
                
                # 1. Hias Bagian Header (Baris ke-1)
                for cell in worksheet[1]:
                    cell.fill = warna_header
                    cell.font = teks_putih
                    cell.alignment = rata_tengah
                
                # 2. Hias Efek Zebra & Rata Tengah Data (Baris ke-2 dst)
                for row_idx in range(2, worksheet.max_row + 1):
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = teks_biasa
                        cell.alignment = rata_tengah
                        
                        # Efek Zebra: Baris genap di Excel (indeks ganjil di perhitungan) diberi latar biru muda
                        if row_idx % 2 == 1:
                            cell.fill = warna_zebra
                
                # 3. Otomatisasi Lebar Kolom (Auto-fit Columns) agar Deskripsi Mengikuti Panjang Teks
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    # Atur lebar kolom dengan toleransi ruang spasi tambahan (+4)
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
            st.download_button(
                label=f"📥 KLIK UNTUK UNDUH FILE ({sheet_aktif}.xlsx)",
                data=buffer.getvalue(),
                file_name=f"Laporan_Cashbook_{sheet_aktif}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # ==================== FITUR RESET TOTAL (HAPUS DATA & HAPUS SHEET) ====================
    st.write("---")
    st.subheader("⚙️ RESET/HAPUS DATA")
    
    if not df_master.empty:
        if st.button(f"⚠️ Hapus 1 Baris Transaksi Terakhir di Tab {sheet_aktif}", use_container_width=True):
            try:
                total_baris_fisik = len(ws_aktif.get_all_values())
                if total_baris_fisik > 1:
                    ws_aktif.delete_rows(total_baris_fisik)
                    perbarui_desain_visual_sheet(ws_aktif)
                    st.success("Baris terakhir di tab ini berhasil dihapus!")
                    st.rerun()
                else:
                    st.warning("Tab sudah kosong, tidak ada data transaksi yang bisa dihapus.")
            except Exception as e:
                st.error(f"Gagal menghapus baris: {e}")
                
    st.write("")
    konfirmasi_reset = st.checkbox("Saya ingin MENGHAPUS SEMUA DATA dan SEMUA TAB BULANAN secara permanen")
    if st.button("🚨 WIPE OUT: HAPUS SEMUA DATA & SEMUA SHEET", type="primary", use_container_width=True, disabled=not konfirmasi_reset):
        try:
            sheet_sementara = f"Mulai_Baru_{rekomendasi_sheet_baru}"
            ws_baru = sh.add_worksheet(title=sheet_sementara, rows="1000", cols="20")
            ws_baru.append_row(["Tanggal", "Hari", "Deskripsi", "Nota", "Debit", "Kredit", "Saldo"])
            ws_baru.append_row([datetime.now().strftime("%Y-%m-%d"), ambil_hari_ini(), "Sistem Direset Total", "-", 0, 0, 0])
            
            for ws in sh.worksheets():
                if ws.title != sheet_sementara:
                    sh.del_worksheet(ws)
            
            ws_baru.update_title(rekomendasi_sheet_baru)
            perbarui_desain_visual_sheet(ws_baru)
            
            st.success("💥 BERHASIL! Semua tab lama dibuang total dan sistem kembali bersih dari nol dengan desain baru!")
            st.rerun()
        except Exception as e:
            st.error(f"Gagal melakukan pembersihan total sheet: {e}")
